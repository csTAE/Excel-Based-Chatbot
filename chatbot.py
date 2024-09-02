import openpyxl
import pandas as pd


workbook = openpyxl.load_workbook('data.xlsx')
sheet = workbook.active


data = pd.DataFrame(sheet.values)
data.columns = data.iloc[0]  
data = data[1:] 

data['Age'] = pd.to_numeric(data['Age'], errors='coerce')
data['Salary'] = pd.to_numeric(data['Salary'], errors='coerce')

def calculate_statistics():
   
    average_age = data['Age'].mean()
    average_salary = data['Salary'].mean()
    most_common_department = data['Department'].mode()[0]
    total_entries = data.shape[0]
    max_salary = data['Salary'].max()
    min_salary = data['Salary'].min()
    oldest_employee = data[data['Age'] == data['Age'].max()]['Name'].values[0]
    youngest_employee = data[data['Age'] == data['Age'].min()]['Name'].values[0]

    stats = (
        f"Average Age: {average_age:.2f}\n"
        f"Average Salary: ${average_salary:.2f}\n"
        f"Most Common Department: {most_common_department}\n"
        f"Total Entries: {total_entries}\n"
        f"Maximum Salary: ${max_salary}\n"
        f"Minimum Salary: ${min_salary}\n"
        f"Oldest Employee: {oldest_employee}\n"
        f"Youngest Employee: {youngest_employee}"
    )
    
    return stats

def calculate_column_average(column_name):
    if column_name in data.columns:
        average = data[column_name].mean()
        return f"Average of column '{column_name}': {average:.2f}"
    else:
        return f"Column '{column_name}' does not exist."

def count_entries():
    return f"Total number of entries: {data.shape[0]}"

def most_demanding_department():
    department_counts = data['Department'].value_counts()
    most_demanding = department_counts.idxmax()
    return f"The most demanding department is: {most_demanding} with {department_counts.max()} employees."

def list_column_names():
    columns = data.columns.tolist()
    return "Column names: " + ", ".join(columns)

def search_data(query, column=None):
    results = []
    if column is not None:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if query.lower() in str(row[column-1]).lower():
                results.append(row)
    else:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if query.lower() in map(lambda x: str(x).lower(), row):
                results.append(row)
    return results

def chatbot():
    print("-------- Hello! I'm an Excel-based chatbot. Ask me anything about the data. ---------")
    
    while True:
        query = input("You: ").strip()
        if query.lower() in ['exit', 'quit', 'bye']:
            print("Chatbot: Goodbye!")
            break

        if query.lower() in ['Hi', 'hii']:
            print("Chatbot:Hi,How can i help.")
        elif query.lower() in ['statistics', 'show me statistics']:
            print("Chatbot:")
            print(calculate_statistics())
        elif query.lower().startswith('average of'):
            column_name = query[len('average of '):].strip()
            print("Chatbot:")
            print(calculate_column_average(column_name))
        elif query.lower() in ['how many entries are there', 'total entries', 'number of entries']:
            print("Chatbot:")
            print(count_entries())
        elif query.lower() in ['most demanding department', ' depademandingrtment']:
            print("Chatbot:")
            print(most_demanding_department())
        elif query.lower() in ['show me the columns names', 'list columns', 'column names']:
            print("Chatbot:")
            print(list_column_names())
        else:
            column_input = input("Which column would you like to search in? (Leave blank for all columns): ").strip()
            column = int(column_input) if column_input.isdigit() else None
            results = search_data(query, column if column else None)
            
            if results:
                for result in results:
                    print(f"Chatbot: {result}")
            else:
                print("Chatbot: No matching records found.")

chatbot()
